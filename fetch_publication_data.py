import requests, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

beamlines = {
    '2-ID':    ['Soft Inelastic X-ray Scattering','SIX', 'Soft X-ray Scattering & Spectroscopy'],
    '3-ID':    ['Hard X-ray Nanoprobe','HXN', 'Imaging'],
    '4-BM':    ['X-ray Fluorescence Microscopy','XFM', 'Imaging'],
    '4-ID':    ['Integrated In situ and Resonant Hard X-ray Studies','ISR', 'Complex Scattering'],
    '5-ID':    ['Submicron Resolution X-ray Spectroscopy','SRX', 'Imaging'],
    '6-BM':    ['Beamline for Materials Measurement','BMM', 'Hard X-ray Scattering & Spectroscopy'],
    '7-BM':    ['Quick x-ray Absorption and Scattering','QAS', 'Hard X-ray Scattering & Spectroscopy'],
    '7-ID-1':  ['Spectroscopy Soft and Tender','SST1', 'Soft X-ray Scattering & Spectroscopy'],
    '7-ID-2':  ['Spectroscopy Soft and Tender 2','SST2', 'Soft X-ray Scattering & Spectroscopy'],
    '8-BM':    ['Tender Energy X-ray Absorption Spectroscopy','TES', 'Imaging'],
    '8-ID':    ['Inner-Shell Spectroscopy','ISS', 'Hard X-ray Scattering & Spectroscopy'],
    '10-ID':   ['Inelastic X-ray Scattering','IXS', 'Complex Scattering'],
    '11-BM':   ['Complex Materials Scattering','CMS', 'Complex Scattering'],
    '11-ID':   ['Coherent Hard X-ray Scattering','CHX', 'Complex Scattering'],
    '12-ID':   ['Soft Matter Interfaces','SMI', 'Complex Scattering'],
    '16-ID':   ['Life Science X-ray Scattering','LiX', 'Structural Biology'],
    '17-BM':   ['X-ray Footprinting of Biological Materials','XFP', 'Structural Biology'],
    '17-ID-1': ['Highly Automated Macromolecular Crystallography','AMX', 'Structural Biology'],
    '17-ID-2': ['Frontier Microfocusing Macromolecular Crystallography','FMX', 'Structural Biology'],
    '18-ID':   ['Full Field X-ray Imaging','FXI', 'Imaging'],
    '19-ID':   ['Biological Microdiffraction Facility','NYX', 'Structural Biology'],
    '21-ID':   ['Electron Spectro-Microscopy','ESM', 'Soft X-ray Scattering & Spectroscopy'],
    '22-IR-1': ['Frontier Synchrotron Infrared Spectroscopy','FIS', 'Soft X-ray Scattering & Spectroscopy'],
    '22-IR-2': ['Magnetospectroscopy, Ellipsometry and Time-Resolved Optical Spectroscopie','MET', 'Soft X-ray Scattering & Spectroscopy'],
    '23-ID-1': ['Coherent Soft X-ray Scattering beamline','CSX', 'Soft X-ray Scattering & Spectroscopy'],
    '23-ID-2': ['In situ and Operando Soft X-ray Spectroscopy','IOS', 'Soft X-ray Scattering & Spectroscopy'],
    '27-ID':   ['High Energy Engineering X-ray Scattering','HEX', 'Hard X-ray Scattering & Spectroscopy'],
    '28-ID-1': ['Pair Distribution Function','PDF', 'Hard X-ray Scattering & Spectroscopy'],
    '28-ID-2': ['X-ray Powder Diffraction','XPD', 'Hard X-ray Scattering & Spectroscopy'],
}

wb = Workbook()
ws = wb.active
ws.append(['port', 'TLA', 'program', 'total', 'high profile', 'years operating',
           '% high profile', 'pubs per year', 'total citations', 'citations per paper'])
c = ws['A2']
ws.freeze_panes = c
for cell in range(20):
    ws[f'{chr(65+cell)}1'].font = Font(bold=True)

for i, port in enumerate(beamlines.keys()):
    tla = beamlines[port][1]
    program = beamlines[port][2]
    print(tla)
    url = f'https://www.bnl.gov/nsls2/beamlines/publications.php?q={port}'
    html = requests.get(url).content
    df_list = pd.read_html(html)
    df = df_list[-1]
    answer = str(df).split('\n')[-1].split()
    years = len(str(df).split('\n')) - 2

    matches=re.findall('Cited (\d+) times', str(html))
    total_citations = sum((int(x) for x in matches))

    ws.append([port, tla, program, int(answer[3]), int(answer[2]), years,
               f'=E{i+3}/D{i+3}',
               f'=D{i+3}/F{i+3}',
               total_citations,
               f'=I{i+3}/D{i+3}',
               ])
    ws[f'G{i+3}'].number_format = '0.00'
    ws[f'H{i+3}'].number_format = '0.00'
    ws[f'J{i+3}'].number_format = '0.00'



ws.column_dimensions['C'].width = 2   *  ws.column_dimensions['D'].width
ws.column_dimensions['E'].width = 1.2 *  ws.column_dimensions['D'].width
ws.column_dimensions['F'].width = 1.3 *  ws.column_dimensions['E'].width
ws.column_dimensions['G'].width = 1.2 *  ws.column_dimensions['F'].width
ws.column_dimensions['H'].width = 1   *  ws.column_dimensions['G'].width
ws.column_dimensions['I'].width = 1   *  ws.column_dimensions['H'].width
ws.column_dimensions['J'].width = 1   *  ws.column_dimensions['H'].width
wb.save("beamline_publication_data.xlsx")
